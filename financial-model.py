#!/usr/bin/env python3
"""
SaaS Financial Model Generator
Creates a comprehensive Excel financial model for a SaaS product including
revenue projections, customer metrics, expenses, and profitability analysis.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.formatting.rule import CellIsRule
from openpyxl.workbook.defined_name import DefinedName


def create_saas_financial_model(filename='saas_financial_model.xlsx'):
    """Create a comprehensive SaaS financial model in Excel."""

    wb = Workbook()

    # Create worksheets
    ws_assumptions = wb.active
    ws_assumptions.title = "Assumptions"
    ws_model = wb.create_sheet("Financial Model")
    ws_dashboard = wb.create_sheet("Dashboard")

    # Define colors and styles
    header_fill = PatternFill(start_color="002E5D", end_color="002E5D", fill_type="solid")
    subheader_fill = PatternFill(start_color="0057B8", end_color="0057B8", fill_type="solid")
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    subheader_font = Font(bold=True, color="FFFFFF", size=11)
    bold_font = Font(bold=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ==================== ASSUMPTIONS SHEET ====================
    ws_assumptions['A1'] = "SaaS Financial Model - Assumptions"
    ws_assumptions['A1'].font = Font(bold=True, size=14)

    # Pricing Assumptions
    ws_assumptions['A3'] = "Pricing Assumptions"
    ws_assumptions['A3'].font = bold_font
    ws_assumptions['A3'].fill = subheader_fill
    ws_assumptions['A3'].font = subheader_font

    assumptions_data = [
        ('Monthly Subscription Price', 50, '$'),
        ('Annual Subscription Price', 500, '$'),
        ('Monthly/Annual Mix - Monthly %', 0.70, '%'),
        ('Monthly/Annual Mix - Annual %', 0.30, '%'),
    ]

    row = 4
    for label, value, format_type in assumptions_data:
        ws_assumptions[f'A{row}'] = label
        ws_assumptions[f'B{row}'] = value
        ws_assumptions[f'B{row}'].fill = input_fill
        if format_type == '$':
            ws_assumptions[f'B{row}'].number_format = '$#,##0'
        elif format_type == '%':
            ws_assumptions[f'B{row}'].number_format = '0%'
        row += 1

    # Customer Assumptions
    row += 1
    ws_assumptions[f'A{row}'] = "Customer Assumptions"
    ws_assumptions[f'A{row}'].font = subheader_font
    ws_assumptions[f'A{row}'].fill = subheader_fill
    row += 1

    customer_assumptions = [
        ('Starting Customers', 100, '#'),
        ('Monthly New Customer Acquisition', 50, '#'),
        ('Monthly Churn Rate', 0.05, '%'),
        ('Customer Acquisition Cost (CAC)', 200, '$'),
    ]

    for label, value, format_type in customer_assumptions:
        ws_assumptions[f'A{row}'] = label
        ws_assumptions[f'B{row}'] = value
        ws_assumptions[f'B{row}'].fill = input_fill
        if format_type == '$':
            ws_assumptions[f'B{row}'].number_format = '$#,##0'
        elif format_type == '%':
            ws_assumptions[f'B{row}'].number_format = '0%'
        elif format_type == '#':
            ws_assumptions[f'B{row}'].number_format = '#,##0'
        row += 1

    # Cost Assumptions
    row += 1
    ws_assumptions[f'A{row}'] = "Cost Assumptions"
    ws_assumptions[f'A{row}'].font = subheader_font
    ws_assumptions[f'A{row}'].fill = subheader_fill
    row += 1

    cost_assumptions = [
        ('COGS per Customer per Month', 10, '$'),
        ('Sales & Marketing % of Revenue', 0.40, '%'),
        ('Fixed Operating Expenses (Monthly)', 50000, '$'),
    ]

    for label, value, format_type in cost_assumptions:
        ws_assumptions[f'A{row}'] = label
        ws_assumptions[f'B{row}'] = value
        ws_assumptions[f'B{row}'].fill = input_fill
        if format_type == '$':
            ws_assumptions[f'B{row}'].number_format = '$#,##0'
        elif format_type == '%':
            ws_assumptions[f'B{row}'].number_format = '0%'
        row += 1

    # Define named ranges for assumptions
    wb.defined_names['monthly_price'] = DefinedName('monthly_price', attr_text="Assumptions!$B$4")
    wb.defined_names['annual_price'] = DefinedName('annual_price', attr_text="Assumptions!$B$5")
    wb.defined_names['monthly_mix'] = DefinedName('monthly_mix', attr_text="Assumptions!$B$6")
    wb.defined_names['annual_mix'] = DefinedName('annual_mix', attr_text="Assumptions!$B$7")
    wb.defined_names['starting_customers'] = DefinedName('starting_customers', attr_text="Assumptions!$B$10")
    wb.defined_names['new_customers_monthly'] = DefinedName('new_customers_monthly', attr_text="Assumptions!$B$11")
    wb.defined_names['churn_rate'] = DefinedName('churn_rate', attr_text="Assumptions!$B$12")
    wb.defined_names['cac'] = DefinedName('cac', attr_text="Assumptions!$B$13")
    wb.defined_names['cogs_per_customer'] = DefinedName('cogs_per_customer', attr_text="Assumptions!$B$16")
    wb.defined_names['sm_percent'] = DefinedName('sm_percent', attr_text="Assumptions!$B$17")
    wb.defined_names['opex_fixed'] = DefinedName('opex_fixed', attr_text="Assumptions!$B$18")

    ws_assumptions.column_dimensions['A'].width = 35
    ws_assumptions.column_dimensions['B'].width = 20

    # ==================== FINANCIAL MODEL SHEET ====================

    # Headers
    ws_model['A1'] = "Month"
    ws_model['A1'].font = header_font
    ws_model['A1'].fill = header_fill

    # Create month headers (0-36)
    for month in range(37):
        col = get_column_letter(month + 2)
        ws_model[f'{col}1'] = month
        ws_model[f'{col}1'].font = header_font
        ws_model[f'{col}1'].fill = header_fill
        ws_model[f'{col}1'].alignment = Alignment(horizontal='center')

    # Section: Customer Metrics
    row = 3
    ws_model[f'A{row}'] = "CUSTOMER METRICS"
    ws_model[f'A{row}'].font = subheader_font
    ws_model[f'A{row}'].fill = subheader_fill
    row += 1

    customer_metrics = [
        'Beginning Customers',
        'New Customers',
        'Churned Customers',
        'Ending Customers'
    ]

    for metric in customer_metrics:
        ws_model[f'A{row}'] = metric
        ws_model[f'A{row}'].font = bold_font
        row += 1

    # Customer formulas
    for month in range(37):
        col = get_column_letter(month + 2)

        # Beginning Customers
        if month == 0:
            ws_model[f'{col}4'] = '=starting_customers'
        else:
            prev_col = get_column_letter(month + 1)
            ws_model[f'{col}4'] = f'={prev_col}7'  # Previous month's ending customers

        # New Customers
        ws_model[f'{col}5'] = '=new_customers_monthly'

        # Churned Customers
        ws_model[f'{col}6'] = f'={col}4*churn_rate'

        # Ending Customers
        ws_model[f'{col}7'] = f'={col}4+{col}5-{col}6'

    # Section: Revenue
    row = 9
    ws_model[f'A{row}'] = "REVENUE"
    ws_model[f'A{row}'].font = subheader_font
    ws_model[f'A{row}'].fill = subheader_fill
    row += 1

    revenue_metrics = [
        'Monthly Subscription Customers',
        'Annual Subscription Customers',
        'MRR (Monthly Subs)',
        'ARR Recognition (Monthly)',
        'Total Monthly Revenue',
        'ARR (Annual Run Rate)'
    ]

    for metric in revenue_metrics:
        ws_model[f'A{row}'] = metric
        ws_model[f'A{row}'].font = bold_font
        row += 1

    # Revenue formulas
    for month in range(37):
        col = get_column_letter(month + 2)

        # Monthly Subscription Customers (70% of ending customers)
        ws_model[f'{col}10'] = f'={col}7*monthly_mix'

        # Annual Subscription Customers (30% of ending customers)
        ws_model[f'{col}11'] = f'={col}7*annual_mix'

        # MRR (Monthly Subs)
        ws_model[f'{col}12'] = f'={col}10*monthly_price'

        # ARR Recognition (Annual subs recognized monthly)
        ws_model[f'{col}13'] = f'={col}11*annual_price/12'

        # Total Monthly Revenue
        ws_model[f'{col}14'] = f'={col}12+{col}13'

        # ARR (Annual Run Rate)
        ws_model[f'{col}15'] = f'={col}14*12'

    # Section: Expenses
    row = 17
    ws_model[f'A{row}'] = "EXPENSES"
    ws_model[f'A{row}'].font = subheader_font
    ws_model[f'A{row}'].fill = subheader_fill
    row += 1

    expense_metrics = [
        'CAC Costs',
        'COGS',
        'Sales & Marketing',
        'Operating Expenses',
        'Total Expenses'
    ]

    for metric in expense_metrics:
        ws_model[f'A{row}'] = metric
        ws_model[f'A{row}'].font = bold_font
        row += 1

    # Expense formulas
    for month in range(37):
        col = get_column_letter(month + 2)

        # CAC Costs
        ws_model[f'{col}18'] = f'={col}5*cac'

        # COGS
        ws_model[f'{col}19'] = f'={col}7*cogs_per_customer'

        # Sales & Marketing
        ws_model[f'{col}20'] = f'={col}14*sm_percent'

        # Operating Expenses
        ws_model[f'{col}21'] = '=opex_fixed'

        # Total Expenses
        ws_model[f'{col}22'] = f'=SUM({col}18:{col}21)'

    # Section: Profitability
    row = 24
    ws_model[f'A{row}'] = "PROFITABILITY"
    ws_model[f'A{row}'].font = subheader_font
    ws_model[f'A{row}'].fill = subheader_fill
    row += 1

    profit_metrics = [
        'Gross Profit',
        'Gross Margin %',
        'Operating Profit',
        'Operating Margin %',
        'Monthly Cash Flow',
        'Cumulative Cash Flow'
    ]

    for metric in profit_metrics:
        ws_model[f'A{row}'] = metric
        ws_model[f'A{row}'].font = bold_font
        row += 1

    # Profitability formulas
    for month in range(37):
        col = get_column_letter(month + 2)

        # Gross Profit (Revenue - COGS)
        ws_model[f'{col}25'] = f'={col}14-{col}19'

        # Gross Margin %
        ws_model[f'{col}26'] = f'=IF({col}14=0,0,{col}25/{col}14)'

        # Operating Profit (Revenue - Total Expenses)
        ws_model[f'{col}27'] = f'={col}14-{col}22'

        # Operating Margin %
        ws_model[f'{col}28'] = f'=IF({col}14=0,0,{col}27/{col}14)'

        # Monthly Cash Flow (same as Operating Profit in this model)
        ws_model[f'{col}29'] = f'={col}27'

        # Cumulative Cash Flow
        if month == 0:
            ws_model[f'{col}30'] = f'={col}29'
        else:
            prev_col = get_column_letter(month + 1)
            ws_model[f'{col}30'] = f'={prev_col}30+{col}29'

    # Section: Key Metrics
    row = 32
    ws_model[f'A{row}'] = "KEY METRICS"
    ws_model[f'A{row}'].font = subheader_font
    ws_model[f'A{row}'].fill = subheader_fill
    row += 1

    key_metrics = [
        'Customer Lifetime Value (LTV)',
        'LTV:CAC Ratio'
    ]

    for metric in key_metrics:
        ws_model[f'A{row}'] = metric
        ws_model[f'A{row}'].font = bold_font
        row += 1

    # Key Metrics formulas
    for month in range(37):
        col = get_column_letter(month + 2)

        # Customer Lifetime Value (Avg revenue per customer / churn rate)
        # Using blended average: 70% pay monthly, 30% pay annual
        ws_model[f'{col}33'] = f'=(monthly_price*monthly_mix + annual_price/12*annual_mix)/churn_rate'

        # LTV:CAC Ratio
        ws_model[f'{col}34'] = f'={col}33/cac'

    # Apply number formatting
    for month in range(37):
        col = get_column_letter(month + 2)

        # Customer counts (no decimals)
        for r in [4, 5, 6, 7, 10, 11]:
            ws_model[f'{col}{r}'].number_format = '#,##0'

        # Currency
        for r in [12, 13, 14, 15, 18, 19, 20, 21, 22, 25, 27, 29, 30, 33]:
            ws_model[f'{col}{r}'].number_format = '$#,##0'

        # Percentages
        for r in [26, 28]:
            ws_model[f'{col}{r}'].number_format = '0.0%'

        # Ratios
        ws_model[f'{col}34'].number_format = '0.0'

    # Conditional formatting for negative cash flow
    negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    negative_font = Font(color="9C0006")

    ws_model.conditional_formatting.add(
        'B29:AL29',
        CellIsRule(operator='lessThan', formula=['0'], fill=negative_fill, font=negative_font)
    )

    ws_model.conditional_formatting.add(
        'B30:AL30',
        CellIsRule(operator='lessThan', formula=['0'], fill=negative_fill, font=negative_font)
    )

    # Adjust column widths
    ws_model.column_dimensions['A'].width = 30
    for col in range(2, 40):
        ws_model.column_dimensions[get_column_letter(col)].width = 12

    # ==================== DASHBOARD SHEET ====================

    ws_dashboard['A1'] = "SaaS Financial Model Dashboard"
    ws_dashboard['A1'].font = Font(bold=True, size=16)

    # Key Milestones Summary
    ws_dashboard['A3'] = "KEY MILESTONES"
    ws_dashboard['A3'].font = subheader_font
    ws_dashboard['A3'].fill = subheader_fill
    ws_dashboard.merge_cells('A3:E3')

    milestone_headers = ['Metric', 'Month 12', 'Month 24', 'Month 36']
    for col_idx, header in enumerate(milestone_headers, start=1):
        cell = ws_dashboard.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = bold_font
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    dashboard_metrics = [
        ('Total Customers', "'Financial Model'!G7", "'Financial Model'!Q7", "'Financial Model'!AK7", '#,##0'),
        ('Monthly Revenue', "'Financial Model'!G14", "'Financial Model'!Q14", "'Financial Model'!AK14", '$#,##0'),
        ('ARR', "'Financial Model'!G15", "'Financial Model'!Q15", "'Financial Model'!AK15", '$#,##0'),
        ('Gross Margin %', "'Financial Model'!G26", "'Financial Model'!Q26", "'Financial Model'!AK26", '0.0%'),
        ('Operating Margin %', "'Financial Model'!G28", "'Financial Model'!Q28", "'Financial Model'!AK28", '0.0%'),
        ('LTV:CAC Ratio', "'Financial Model'!G34", "'Financial Model'!Q34", "'Financial Model'!AK34", '0.0'),
        ('Cumulative Cash Flow', "'Financial Model'!G30", "'Financial Model'!Q30", "'Financial Model'!AK30", '$#,##0'),
    ]

    for idx, (metric_name, m12_ref, m24_ref, m36_ref, num_format) in enumerate(dashboard_metrics, start=5):
        ws_dashboard[f'A{idx}'] = metric_name
        ws_dashboard[f'A{idx}'].font = bold_font

        ws_dashboard[f'B{idx}'] = f'={m12_ref}'
        ws_dashboard[f'B{idx}'].number_format = num_format

        ws_dashboard[f'C{idx}'] = f'={m24_ref}'
        ws_dashboard[f'C{idx}'].number_format = num_format

        ws_dashboard[f'D{idx}'] = f'={m36_ref}'
        ws_dashboard[f'D{idx}'].number_format = num_format

    # Business Health Indicators
    ws_dashboard['A14'] = "BUSINESS HEALTH INDICATORS"
    ws_dashboard['A14'].font = subheader_font
    ws_dashboard['A14'].fill = subheader_fill
    ws_dashboard.merge_cells('A14:C14')

    ws_dashboard['A15'] = "Average Gross Margin (36 months)"
    ws_dashboard['A15'].font = bold_font
    ws_dashboard['B15'] = "=AVERAGE('Financial Model'!B26:AK26)"
    ws_dashboard['B15'].number_format = '0.0%'

    ws_dashboard['A16'] = "Average LTV:CAC Ratio (36 months)"
    ws_dashboard['A16'].font = bold_font
    ws_dashboard['B16'] = "=AVERAGE('Financial Model'!B34:AK34)"
    ws_dashboard['B16'].number_format = '0.0'

    ws_dashboard['A17'] = "Months to Profitability"
    ws_dashboard['A17'].font = bold_font
    ws_dashboard['B17'] = "=IFERROR(MATCH(TRUE,'Financial Model'!B29:AK29>0,0),\"Not Yet\")"

    ws_dashboard['A18'] = "Months to Positive Cumulative Cash Flow"
    ws_dashboard['A18'].font = bold_font
    ws_dashboard['B18'] = "=IFERROR(MATCH(TRUE,'Financial Model'!B30:AK30>0,0),\"Not Yet\")"

    # Chart 1: Revenue Growth
    chart1 = LineChart()
    chart1.title = "Revenue Growth Over Time"
    chart1.style = 10
    chart1.y_axis.title = "Revenue ($)"
    chart1.x_axis.title = "Month"
    chart1.height = 10
    chart1.width = 20

    data = Reference(ws_model, min_col=2, min_row=14, max_col=38, max_row=14)
    cats = Reference(ws_model, min_col=2, min_row=1, max_col=38)
    chart1.add_data(data, titles_from_data=False)
    chart1.set_categories(cats)
    chart1.series[0].graphicalProperties.line.width = 2.5 * 12700  # in EMUs

    ws_dashboard.add_chart(chart1, "F3")

    # Chart 2: Customer Growth
    chart2 = LineChart()
    chart2.title = "Customer Growth Over Time"
    chart2.style = 11
    chart2.y_axis.title = "Customers"
    chart2.x_axis.title = "Month"
    chart2.height = 10
    chart2.width = 20

    data = Reference(ws_model, min_col=2, min_row=7, max_col=38, max_row=7)
    cats = Reference(ws_model, min_col=2, min_row=1, max_col=38)
    chart2.add_data(data, titles_from_data=False)
    chart2.set_categories(cats)
    chart2.series[0].graphicalProperties.line.width = 2.5 * 12700

    ws_dashboard.add_chart(chart2, "F20")

    # Chart 3: Profitability
    chart3 = LineChart()
    chart3.title = "Cumulative Cash Flow"
    chart3.style = 12
    chart3.y_axis.title = "Cash Flow ($)"
    chart3.x_axis.title = "Month"
    chart3.height = 10
    chart3.width = 20

    data = Reference(ws_model, min_col=2, min_row=30, max_col=38, max_row=30)
    cats = Reference(ws_model, min_col=2, min_row=1, max_col=38)
    chart3.add_data(data, titles_from_data=False)
    chart3.set_categories(cats)
    chart3.series[0].graphicalProperties.line.width = 2.5 * 12700

    ws_dashboard.add_chart(chart3, "R3")

    # Adjust column widths
    ws_dashboard.column_dimensions['A'].width = 35
    ws_dashboard.column_dimensions['B'].width = 18
    ws_dashboard.column_dimensions['C'].width = 18
    ws_dashboard.column_dimensions['D'].width = 18

    # Save the workbook
    wb.save(filename)
    print(f"Financial model created successfully: {filename}")
    print(f"The model includes:")
    print(f"  - Assumptions sheet with editable inputs")
    print(f"  - Financial Model sheet with 36-month projections")
    print(f"  - Dashboard with key metrics and charts")
    print(f"\nKey features:")
    print(f"  - Customer metrics (acquisition, churn, LTV)")
    print(f"  - Revenue calculations (MRR, ARR)")
    print(f"  - Expense tracking (CAC, COGS, S&M, OpEx)")
    print(f"  - Profitability analysis (gross margin, operating margin)")
    print(f"  - Cash flow projections")
    print(f"  - Conditional formatting for negative cash flow")


if __name__ == "__main__":
    create_saas_financial_model()
